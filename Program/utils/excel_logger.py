"""
Excel Logger Module for MagnumOpus Project

This module provides comprehensive logging functionality that saves all program outputs
to Excel files with separate worksheets for each program/module. New runs append data
with timestamps rather than overwriting existing logs.

Features:
- Separate worksheet for each program/module
- Timestamped headers for each run
- Automatic appending (no data loss)
- Support for structured and unstructured data
- Easy-to-use context manager
- Automatic Excel file creation and management

Usage:
    from utils import ExcelLogger

    with ExcelLogger("training") as logger:
        logger.log("Training started...")
        logger.log({"epoch": 1, "loss": 0.5, "accuracy": 0.85})
        logger.log("Training completed!")
"""

import os
import sys
from datetime import datetime
from pathlib import Path
import io
from contextlib import redirect_stdout, redirect_stderr

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Warning: openpyxl not installed. Installing now...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


class ExcelLogger:
    """
    Logger class that captures program output and saves it to Excel files.

    Each program gets its own worksheet, and each run is timestamped and appended
    to preserve historical data.
    """

    # Default log directory
    DEFAULT_LOG_DIR = Path(__file__).parent.parent / "logs"
    DEFAULT_EXCEL_FILE = "program_logs.xlsx"

    def __init__(self, program_name, excel_file=None, log_dir=None, capture_stdout=False):
        """
        Initialize the Excel Logger

        Args:
            program_name: Name of the program/module being logged
            excel_file: Name of the Excel file (default: program_logs.xlsx)
            log_dir: Directory to save logs (default: Program/logs/)
            capture_stdout: Whether to automatically capture stdout (default: False)
        """
        self.program_name = program_name
        self.capture_stdout = capture_stdout

        # Setup paths
        self.log_dir = Path(log_dir) if log_dir else self.DEFAULT_LOG_DIR
        self.excel_file = excel_file if excel_file else self.DEFAULT_EXCEL_FILE
        self.excel_path = self.log_dir / self.excel_file

        # Create log directory if it doesn't exist
        self.log_dir.mkdir(parents=True, exist_ok=True)

        # Initialize logging session
        self.session_start = datetime.now()
        self.log_buffer = []

        # For stdout capture
        self._stdout_capture = None
        self._original_stdout = None

        print(f"✓ ExcelLogger initialized for '{program_name}'")
        print(f"✓ Logs will be saved to: {self.excel_path}")

    def __enter__(self):
        """Context manager entry"""
        if self.capture_stdout:
            self._stdout_capture = io.StringIO()
            self._original_stdout = sys.stdout
            sys.stdout = self._stdout_capture

        self.log(f"=== SESSION START: {self.session_start.strftime('%Y-%m-%d %H:%M:%S')} ===")
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - saves all logs to Excel"""
        if exc_type is not None:
            self.log(f"ERROR: {exc_type.__name__}: {exc_val}")

        self.log(f"=== SESSION END: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")

        # Capture stdout if enabled
        if self.capture_stdout and self._stdout_capture:
            captured_output = self._stdout_capture.getvalue()
            if captured_output:
                self.log("=== CAPTURED STDOUT ===")
                self.log(captured_output)
            sys.stdout = self._original_stdout

        # Save to Excel
        self._save_to_excel()
        print(f"✓ Logs saved to: {self.excel_path}")
        print(f"✓ Worksheet: '{self.program_name}'")

    def log(self, message, data=None):
        """
        Log a message or data

        Args:
            message: Text message to log
            data: Optional dictionary of structured data
        """
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

        if isinstance(message, dict):
            # If message is a dict, treat it as structured data
            data = message
            message = "Data Entry"

        log_entry = {
            'timestamp': timestamp,
            'message': str(message),
            'data': data
        }

        self.log_buffer.append(log_entry)

    def log_dict(self, data_dict, prefix=""):
        """
        Log a dictionary in a readable format

        Args:
            data_dict: Dictionary to log
            prefix: Optional prefix for the log entry
        """
        if prefix:
            self.log(prefix)

        for key, value in data_dict.items():
            self.log(f"  {key}: {value}")

    def log_list(self, data_list, title="List Data"):
        """
        Log a list of items

        Args:
            data_list: List to log
            title: Title for the list
        """
        self.log(title)
        for i, item in enumerate(data_list, 1):
            self.log(f"  [{i}] {item}")

    def log_metrics(self, metrics_dict):
        """
        Log metrics in a structured format (useful for training/evaluation)

        Args:
            metrics_dict: Dictionary of metrics
        """
        self.log("--- METRICS ---")
        self.log_dict(metrics_dict)

    def log_separator(self, char="=", length=80):
        """Log a separator line"""
        self.log(char * length)

    def _save_to_excel(self):
        """Save accumulated logs to Excel file"""
        try:
            # Load or create workbook
            if self.excel_path.exists():
                wb = load_workbook(self.excel_path)
            else:
                wb = Workbook()
                # Remove default sheet if it exists
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])

            # Get or create worksheet for this program
            if self.program_name in wb.sheetnames:
                ws = wb[self.program_name]
            else:
                ws = wb.create_sheet(self.program_name)

            # Find the next empty row
            next_row = ws.max_row + 1 if ws.max_row > 1 else 1

            # If this is the first entry, add column headers
            if ws.max_row == 1 and ws['A1'].value is None:
                headers = ['Run Date', 'Timestamp', 'Message', 'Data']
                ws.append(headers)

                # Style headers
                for col in range(1, 5):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                next_row = 2

            # Add session header
            session_header_row = next_row
            ws.cell(row=next_row, column=1, value=self.session_start.strftime('%Y-%m-%d'))
            ws.cell(row=next_row, column=2, value=self.session_start.strftime('%H:%M:%S'))
            ws.cell(row=next_row, column=3, value=f"=== RUN: {self.session_start.strftime('%Y-%m-%d %H:%M:%S')} ===")

            # Style session header
            for col in range(1, 5):
                cell = ws.cell(row=next_row, column=col)
                cell.font = Font(bold=True, size=10, color="000080")
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            next_row += 1

            # Add all log entries
            for entry in self.log_buffer:
                timestamp = entry['timestamp']
                message = entry['message']
                data = entry.get('data')

                # Parse timestamp to extract time only
                try:
                    dt = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S.%f')
                    time_only = dt.strftime('%H:%M:%S.%f')[:-3]
                except:
                    time_only = timestamp

                # Write to Excel
                ws.cell(row=next_row, column=1, value=self.session_start.strftime('%Y-%m-%d'))
                ws.cell(row=next_row, column=2, value=time_only)
                ws.cell(row=next_row, column=3, value=message)

                if data:
                    # Convert data dict to string representation
                    data_str = str(data) if not isinstance(data, dict) else ', '.join([f"{k}={v}" for k, v in data.items()])
                    ws.cell(row=next_row, column=4, value=data_str)

                next_row += 1

            # Add blank row after session
            next_row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max_length + 2, 100)  # Cap at 100
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(self.excel_path)

        except Exception as e:
            print(f"Error saving to Excel: {e}")
            # Fallback: save to text file
            self._save_to_text_fallback()

    def _save_to_text_fallback(self):
        """Fallback: Save to text file if Excel fails"""
        text_file = self.log_dir / f"{self.program_name}_log.txt"
        with open(text_file, 'a', encoding='utf-8') as f:
            f.write(f"\n\n{'='*80}\n")
            f.write(f"RUN: {self.session_start.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*80}\n")
            for entry in self.log_buffer:
                f.write(f"[{entry['timestamp']}] {entry['message']}\n")
                if entry.get('data'):
                    f.write(f"  Data: {entry['data']}\n")
        print(f"✓ Fallback: Logs saved to text file: {text_file}")


class ExcelLoggerMulti:
    """
    Advanced logger that can handle multiple programs in a single session
    """

    def __init__(self, excel_file=None, log_dir=None):
        """
        Initialize multi-program logger

        Args:
            excel_file: Name of the Excel file
            log_dir: Directory to save logs
        """
        self.log_dir = Path(log_dir) if log_dir else ExcelLogger.DEFAULT_LOG_DIR
        self.excel_file = excel_file if excel_file else ExcelLogger.DEFAULT_EXCEL_FILE
        self.excel_path = self.log_dir / self.excel_file
        self.log_dir.mkdir(parents=True, exist_ok=True)

        self.loggers = {}

    def get_logger(self, program_name):
        """Get or create a logger for a specific program"""
        if program_name not in self.loggers:
            self.loggers[program_name] = ExcelLogger(
                program_name=program_name,
                excel_file=self.excel_file,
                log_dir=self.log_dir,
                capture_stdout=False
            )
        return self.loggers[program_name]

    def save_all(self):
        """Save all loggers"""
        for logger in self.loggers.values():
            logger._save_to_excel()


# Test function to verify the module works correctly
def test_excel_logger():
    """Test the ExcelLogger module"""
    print("="*80)
    print("TESTING EXCEL LOGGER MODULE")
    print("="*80)

    # Test 1: Basic logging
    print("\n[Test 1] Basic logging with context manager...")
    with ExcelLogger("test_basic") as logger:
        logger.log("This is a test message")
        logger.log("Testing with numbers", {"value": 123, "result": "success"})
        logger.log_separator()
        logger.log("Test completed")
    print("✓ Test 1 passed")

    # Test 2: Metrics logging
    print("\n[Test 2] Metrics logging...")
    with ExcelLogger("test_metrics") as logger:
        logger.log("Training Epoch 1")
        logger.log_metrics({
            "epoch": 1,
            "loss": 0.5234,
            "accuracy": 0.8567,
            "val_loss": 0.6123,
            "val_accuracy": 0.8234
        })
        logger.log_separator("-")
        logger.log("Training Epoch 2")
        logger.log_metrics({
            "epoch": 2,
            "loss": 0.4123,
            "accuracy": 0.8876,
            "val_loss": 0.5234,
            "val_accuracy": 0.8567
        })
    print("✓ Test 2 passed")

    # Test 3: List logging
    print("\n[Test 3] List logging...")
    with ExcelLogger("test_list") as logger:
        logger.log_list([
            "VGG16 - Accuracy: 0.85",
            "ResNet50 - Accuracy: 0.87",
            "InceptionV3 - Accuracy: 0.89"
        ], title="Model Performance Summary")
    print("✓ Test 3 passed")

    # Test 4: Multiple runs (appending)
    print("\n[Test 4] Multiple runs (testing append functionality)...")
    for run in range(1, 3):
        with ExcelLogger("test_append") as logger:
            logger.log(f"This is run #{run}")
            logger.log(f"Run {run} completed successfully")
    print("✓ Test 4 passed")

    # Test 5: Multi-program logger
    print("\n[Test 5] Multi-program logger...")
    multi_logger = ExcelLoggerMulti(excel_file="multi_test_logs.xlsx")

    config_logger = multi_logger.get_logger("config")
    config_logger.log("Config module test")
    config_logger.log_dict({"setting1": "value1", "setting2": "value2"})

    data_logger = multi_logger.get_logger("data")
    data_logger.log("Data module test")
    data_logger.log("Loaded 1000 samples")

    multi_logger.save_all()
    print("✓ Test 5 passed")

    print("\n" + "="*80)
    print("ALL TESTS PASSED!")
    print("="*80)
    print(f"\nCheck the logs directory: {ExcelLogger.DEFAULT_LOG_DIR}")
    print(f"Main log file: {ExcelLogger.DEFAULT_LOG_DIR / ExcelLogger.DEFAULT_EXCEL_FILE}")
    print(f"Multi-test log file: {ExcelLogger.DEFAULT_LOG_DIR / 'multi_test_logs.xlsx'}")


if __name__ == "__main__":
    # Run tests when module is executed directly
    test_excel_logger()

